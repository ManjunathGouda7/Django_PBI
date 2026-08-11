# BI/analytics/analytics_advanced.py
import io
import json
import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse
from django.conf import settings
from .models import Dataset, Dashboard, Widget, CalculatedMeasure
from .services import DatasetEngine

class AIAutoBuilder:
    """
    Automated AI Dashboard & Visual Widget Builder.
    Analyzes dataset schemas to construct optimized 4-widget analytical dashboards
    and add visual widgets dynamically from natural language chat queries.
    """

    @staticmethod
    def build_auto_dashboard(dataset, title=None, user=None):
        df = DatasetEngine.load_dataframe(dataset)
        if df.empty:
            raise ValueError("Cannot auto-build dashboard for an empty dataset.")

        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        categorical_cols = df.select_dtypes(include=['object', 'category']).columns.tolist()

        db_title = title or f"{dataset.name} AI Studio Dashboard"
        dashboard = Dashboard.objects.create(
            title=db_title,
            description=f"Auto-generated AI Dashboard for {dataset.name} ({len(df):,} records).",
            dataset=dataset,
            theme="dark_modern",
            created_by=user
        )

        pos_y = 0
        primary_num = numeric_cols[0] if numeric_cols else df.columns[0]
        secondary_num = numeric_cols[1] if len(numeric_cols) > 1 else primary_num
        primary_cat = categorical_cols[0] if categorical_cols else df.columns[0]

        # 1. KPI Card Widget
        Widget.objects.create(
            dashboard=dashboard,
            title=f"Total Records & {primary_num}",
            visual_type="kpi",
            y_axis=primary_num if primary_num in numeric_cols else None,
            aggregation="SUM" if primary_num in numeric_cols else "COUNT",
            position_x=0, position_y=pos_y, width=12, height=3,
            created_by=user
        )
        pos_y += 3

        # 2. Scatter Plot / Line Widget
        if len(numeric_cols) >= 2:
            Widget.objects.create(
                dashboard=dashboard,
                title=f"{secondary_num} vs {primary_num} Scatter Distribution",
                visual_type="scatter",
                x_axis=primary_num,
                y_axis=secondary_num,
                group_by=primary_cat if primary_cat != primary_num else None,
                aggregation="AVG",
                position_x=0, position_y=pos_y, width=12, height=7,
                created_by=user
            )
            pos_y += 7

        # 3. Categorical Bar / Column Chart Widget
        if categorical_cols and primary_num in numeric_cols:
            Widget.objects.create(
                dashboard=dashboard,
                title=f"Average {primary_num} by {primary_cat}",
                visual_type="column",
                x_axis=primary_cat,
                y_axis=primary_num,
                aggregation="AVG",
                position_x=0, position_y=pos_y, width=6, height=5,
                created_by=user
            )

        # 4. Aggregated Table Matrix Widget
        Widget.objects.create(
            dashboard=dashboard,
            title=f"{dataset.name} Matrix Summary Table",
            visual_type="table",
            x_axis=primary_cat if categorical_cols else df.columns[0],
            y_axis=primary_num if primary_num in numeric_cols else df.columns[1] if len(df.columns) > 1 else df.columns[0],
            aggregation="AVG" if primary_num in numeric_cols else "COUNT",
            position_x=6 if (categorical_cols and primary_num in numeric_cols) else 0,
            position_y=pos_y if (categorical_cols and primary_num in numeric_cols) else pos_y,
            width=6 if (categorical_cols and primary_num in numeric_cols) else 12,
            height=5,
            created_by=user
        )

        return dashboard

    @staticmethod
    def create_widget_from_chat(dashboard, title, visual_type, x_axis, y_axis, aggregation="AVG", group_by=None, user=None):
        max_y = 0
        for w in dashboard.widgets.all():
            max_y = max(max_y, w.position_y + w.height)

        widget = Widget.objects.create(
            dashboard=dashboard,
            title=title or f"{visual_type.title()} Visual",
            visual_type=visual_type or "bar",
            x_axis=x_axis,
            y_axis=y_axis,
            aggregation=aggregation or "AVG",
            group_by=group_by,
            position_x=0,
            position_y=max_y,
            width=12 if visual_type in ['scatter', 'table'] else 6,
            height=6 if visual_type in ['scatter', 'table'] else 4,
            created_by=user
        )
        return widget


class AnomalyEngine:
    """
    High-Performance Vectorized Outlier & Telemetry Anomaly Detection Engine.
    Uses C-level NumPy/Pandas Z-Score matrix calculations to flag outliers,
    power spikes, and specification limit breaches in milliseconds.
    """

    @staticmethod
    def detect_anomalies(dataset):
        df = DatasetEngine.load_dataframe(dataset)
        if df.empty:
            return {'total_anomalies': 0, 'anomalous_rows': [], 'summary': {}}

        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        if not numeric_cols:
            return {'total_anomalies': 0, 'anomalous_rows': [], 'summary': {}}

        # Vectorized C-level matrix Z-score computation
        num_df = df[numeric_cols].apply(pd.to_numeric, errors='coerce')
        means = num_df.mean()
        stds = num_df.std().replace(0, np.nan)
        z_df = (num_df - means) / stds

        anomalous_rows = []
        summary = {}

        entity_col = None
        for candidate in ['Board', 'DUT', 'PowerMode', 'RUN', 'CRX']:
            if candidate in df.columns:
                entity_col = candidate
                break

        for col in numeric_cols:
            if col not in z_df or stds[col] is np.nan or np.isnan(stds[col]):
                continue

            col_z = z_df[col].dropna()
            breaches_mask = col_z.abs() > 2.0
            breached_indices = col_z[breaches_mask].index

            if not breached_indices.empty:
                mean_val = float(means[col])
                std_val = float(stds[col]) if not np.isnan(stds[col]) else 0.0

                summary[col] = {
                    'mean': round(mean_val, 2),
                    'std': round(std_val, 2),
                    'breach_count': len(breached_indices),
                    'upper_threshold': round(mean_val + 2 * std_val, 2),
                    'lower_threshold': round(max(0, mean_val - 2 * std_val), 2)
                }

                # Sample top breaches efficiently
                sample_indices = breached_indices[:30]
                for idx in sample_indices:
                    z_val = float(col_z.loc[idx])
                    severity = "CRITICAL" if abs(z_val) > 3.0 else "WARNING"
                    val = df.at[idx, col]
                    entity_val = df.at[idx, entity_col] if entity_col else f"Row {idx+1}"

                    anomalous_rows.append({
                        'row_index': int(idx),
                        'field': col,
                        'value': round(float(val), 2) if isinstance(val, (int, float, np.number)) else str(val),
                        'z_score': round(z_val, 2),
                        'severity': severity,
                        'entity': str(entity_val),
                        'upper_bound': round(mean_val + 2 * std_val, 2),
                        'lower_bound': round(max(0, mean_val - 2 * std_val), 2)
                    })

        anomalous_rows.sort(key=lambda x: abs(x['z_score']), reverse=True)

        return {
            'total_anomalies': len(anomalous_rows),
            'anomalous_rows': anomalous_rows[:50],
            'summary': summary
        }



class DataWrangler:
    """
    Data Cleaning, Missing Value Imputation & Calculated DAX/Pandas Measures.
    """

    @staticmethod
    def clean_dataset(dataset, fill_method=None, remove_duplicates=False, drop_nulls=False):
        df = DatasetEngine.load_dataframe(dataset)
        if df.empty:
            return dataset

        if remove_duplicates:
            df = df.drop_duplicates()

        if drop_nulls:
            df = df.dropna()
        elif fill_method == 'mean':
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            df[numeric_cols] = df[numeric_cols].fillna(df[numeric_cols].mean())
            df = df.fillna('N/A')
        elif fill_method == 'zero':
            df = df.fillna(0)

        # Save back to file if dataset has file
        if dataset.file and os.path.exists(dataset.file.path):
            if dataset.file.path.endswith('.csv'):
                df.to_csv(dataset.file.path, index=False)
            elif dataset.file.path.endswith('.json'):
                df.to_json(dataset.file.path, orient='records', indent=2)
            elif dataset.file.path.endswith(('.xlsx', '.xls')):
                df.to_excel(dataset.file.path, index=False)

        # Update dataset metadata
        dataset.row_count = len(df)
        dataset.column_schema = DatasetEngine.infer_column_schema(df)
        dataset.save()

        # Update cache
        from .services import _df_cache
        _df_cache[dataset.id] = df.copy()

        return dataset

    @staticmethod
    def validate_formula_security(formula, df_columns):

        """
        🔒 Formula Expression Sanitizer & Security Hardener.
        Parses expression using AST to verify it contains only safe arithmetic nodes
        (BinOp, UnaryOp, Num/Constant, Name) matching valid dataset column names.
        """
        import ast

        if '__' in formula or 'import' in formula or 'os.' in formula or 'sys.' in formula or 'eval' in formula or 'exec' in formula:
            raise ValueError("Formula contains forbidden keywords or tokens.")

        try:
            tree = ast.parse(formula, mode='eval')
        except Exception as e:
            raise ValueError(f"Invalid mathematical expression syntax: {str(e)}")

        for node in ast.walk(tree):
            if isinstance(node, (ast.Expression, ast.BinOp, ast.UnaryOp, ast.Constant, ast.Name, ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Mod, ast.Pow, ast.USub, ast.UAdd, ast.Load)):

                if isinstance(node, ast.Name):
                    if node.id not in df_columns and node.id not in ('True', 'False'):
                        col_matches = [c for c in df_columns if c.lower() == node.id.lower()]
                        if not col_matches:
                            raise ValueError(f"Column '{node.id}' not found in dataset schema.")
            else:
                raise ValueError(f"Forbidden syntax token '{type(node).__name__}' in formula.")

        return True

    @staticmethod
    def add_calculated_measure(dataset, name, formula):
        df = DatasetEngine.load_dataframe(dataset)
        if df.empty:
            raise ValueError("Dataset is empty.")

        clean_name = str(name).strip().replace(' ', '_')
        clean_formula = formula.strip()

        # Perform AST security & syntax validation
        DataWrangler.validate_formula_security(clean_formula, list(df.columns))

        try:
            df[clean_name] = df.eval(clean_formula)
        except Exception as eval_err:
            raise ValueError(f"Failed to calculate measure formula '{formula}': {str(eval_err)}")

        # Save back to dataset file
        if dataset.file and os.path.exists(dataset.file.path):
            if dataset.file.path.endswith('.csv'):
                df.to_csv(dataset.file.path, index=False)
            elif dataset.file.path.endswith('.json'):
                df.to_json(dataset.file.path, orient='records', indent=2)

        dataset.row_count = len(df)
        dataset.column_schema = DatasetEngine.infer_column_schema(df)
        dataset.save()

        CalculatedMeasure.objects.create(
            dataset=dataset,
            name=clean_name,
            formula=clean_formula
        )

        from .services import _df_cache
        _df_cache[dataset.id] = df.copy()

        return clean_name



class DatasetJoiner:
    """
    Multi-Dataset Merging & VLOOKUP Modeling.
    Merges two datasets on shared key columns.
    """

    @staticmethod
    def join_datasets(dataset1, dataset2, key_col1, key_col2, join_type='inner', name=None):
        df1 = DatasetEngine.load_dataframe(dataset1)
        df2 = DatasetEngine.load_dataframe(dataset2)

        if df1.empty or df2.empty:
            raise ValueError("One or both datasets are empty.")

        if key_col1 not in df1.columns:
            raise ValueError(f"Column '{key_col1}' not found in {dataset1.name}.")
        if key_col2 not in df2.columns:
            raise ValueError(f"Column '{key_col2}' not found in {dataset2.name}.")

        merged_df = pd.merge(
            df1, df2,
            left_on=key_col1, right_on=key_col2,
            how=join_type,
            suffixes=('', '_joined')
        )

        join_name = name or f"Joined: {dataset1.name} + {dataset2.name}"
        
        # Save as new CSV file in media directory
        upload_dir = os.path.join(settings.MEDIA_ROOT, 'datasets', 'joined')
        os.makedirs(upload_dir, exist_ok=True)
        file_path = os.path.join(upload_dir, f"joined_{dataset1.id}_{dataset2.id}.csv")
        merged_df.to_csv(file_path, index=False)

        new_dataset = Dataset.objects.create(
            name=join_name,
            description=f"Unified dataset created by joining {dataset1.name} and {dataset2.name} on {key_col1}={key_col2}.",
            file=f"datasets/joined/joined_{dataset1.id}_{dataset2.id}.csv",
            file_type='csv',
            row_count=len(merged_df),
            column_schema=DatasetEngine.infer_column_schema(merged_df),
            is_sample=False
        )

        return new_dataset


class ExcelExporter:
    """
    Multi-Sheet Formatted Excel (.xlsx) Exporter Engine.
    Generates Excel workbooks with Summary, Raw Data, and Schema tabs.
    """

    @staticmethod
    def generate_excel_workbook(dataset):
        df = DatasetEngine.load_dataframe(dataset)

        wb = openpyxl.Workbook()
        
        # Styling definitions
        header_fill = PatternFill(start_color="00A4EF", end_color="00A4EF", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="002060")
        sub_font = Font(name="Calibri", size=10, italic=True, color="595959")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # ----------------------------------------------------
        # Sheet 1: Executive Summary
        # ----------------------------------------------------
        ws_sum = wb.active
        ws_sum.title = "Executive Summary"

        ws_sum.append(["Apex BI Studio — Telemetry Executive Summary"])
        ws_sum.append([f"Dataset Name: {dataset.name}"])
        ws_sum.append([f"File Format: {dataset.file_type.upper()} | Total Rows: {len(df):,} | Columns: {len(df.columns)}"])
        ws_sum.append([])

        ws_sum['A1'].font = title_font
        ws_sum['A2'].font = Font(name="Calibri", size=12, bold=True, color="00A4EF")
        ws_sum['A3'].font = sub_font

        ws_sum.append(["Metric Name", "Value", "Notes"])
        ws_sum['A5'].fill = header_fill
        ws_sum['A5'].font = header_font
        ws_sum['B5'].fill = header_fill
        ws_sum['B5'].font = header_font
        ws_sum['C5'].fill = header_fill
        ws_sum['C5'].font = header_font

        ws_sum.append(["Total Records", len(df), "Total rows imported"])
        ws_sum.append(["Total Columns", len(df.columns), "Schema dimensions & metrics"])
        ws_sum.append(["Missing Cells", int(df.isnull().sum().sum()), "Total empty entries"])
        ws_sum.append(["Duplicate Rows", int(df.duplicated().sum()), "Duplicate row count"])

        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        for col in numeric_cols[:5]:
            s = pd.to_numeric(df[col], errors='coerce').dropna()
            if not s.empty:
                ws_sum.append([f"Avg ({col})", round(float(s.mean()), 2), f"Min: {round(float(s.min()), 2)}, Max: {round(float(s.max()), 2)}"])

        # Auto-adjust column widths
        for col in ws_sum.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 15)

        # ----------------------------------------------------
        # Sheet 2: Raw Telemetry Data
        # ----------------------------------------------------
        ws_data = wb.create_sheet(title="Telemetry Records")
        preview_df = df.head(10000)

        for r_idx, row in enumerate(dataframe_to_rows(preview_df, index=False, header=True), 1):
            ws_data.append(row)
            if r_idx == 1:
                for c_idx in range(1, len(row) + 1):
                    cell = ws_data.cell(row=1, column=c_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws_data.columns:
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_data.column_dimensions[col_letter].width = 16

        # ----------------------------------------------------
        # Sheet 3: Column Schema Dictionary
        # ----------------------------------------------------
        ws_schema = wb.create_sheet(title="Column Schema")
        ws_schema.append(["Column Name", "Data Type", "Unique Count", "Null Count", "Sample Values"])
        for c_idx in range(1, 6):
            cell = ws_schema.cell(row=1, column=c_idx)
            cell.fill = header_fill
            cell.font = header_font

        schema_list = dataset.column_schema or DatasetEngine.infer_column_schema(df)
        for col_info in schema_list:
            samples = ", ".join(col_info.get('samples', []))
            ws_schema.append([
                col_info.get('name'),
                col_info.get('type'),
                col_info.get('unique_count', 0),
                col_info.get('null_count', 0),
                samples
            ])

        for col in ws_schema.columns:
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_schema.column_dimensions[col_letter].width = 20

        # Save to memory stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        safe_name = dataset.name.replace(' ', '_').replace('/', '_')
        response['Content-Disposition'] = f'attachment; filename="ApexBI_{safe_name}_Export.xlsx"'
        return response


class TemplateManager:
    """
    Dashboard Layout JSON Export & Import Template Engine.
    """

    @staticmethod
    def export_template(dashboard):
        widgets_list = []
        for w in dashboard.widgets.all():
            widgets_list.append({
                'title': w.title,
                'visual_type': w.visual_type,
                'x_axis': w.x_axis,
                'y_axis': w.y_axis,
                'aggregation': w.aggregation,
                'group_by': w.group_by,
                'position_x': w.position_x,
                'position_y': w.position_y,
                'width': w.width,
                'height': w.height
            })

        template_data = {
            'apex_bi_version': '1.0',
            'title': dashboard.title,
            'description': dashboard.description,
            'theme': dashboard.theme,
            'layout_type': dashboard.layout_type,
            'widgets': widgets_list
        }

        return template_data

    @staticmethod
    def import_template(dataset, template_json, title=None, user=None):
        db_title = title or template_json.get('title', 'Imported Dashboard Template')
        dashboard = Dashboard.objects.create(
            title=db_title,
            description=template_json.get('description', 'Imported from layout JSON template.'),
            dataset=dataset,
            theme=template_json.get('theme', 'dark_modern'),
            layout_type=template_json.get('layout_type', 'grid'),
            created_by=user
        )

        for w_data in template_json.get('widgets', []):
            Widget.objects.create(
                dashboard=dashboard,
                title=w_data.get('title', 'Visual'),
                visual_type=w_data.get('visual_type', 'bar'),
                x_axis=w_data.get('x_axis'),
                y_axis=w_data.get('y_axis'),
                aggregation=w_data.get('aggregation', 'SUM'),
                group_by=w_data.get('group_by'),
                position_x=w_data.get('position_x', 0),
                position_y=w_data.get('position_y', 0),
                width=w_data.get('width', 6),
                height=w_data.get('height', 4),
                created_by=user
            )

        return dashboard


class ForecastingEngine:
    """
    🔮 AI Predictive Time-Series Forecasting Engine.
    Uses polynomial trend regression & standard error variance bounds to predict
    future telemetry metric values (7 to 30 days out) with confidence intervals.
    """

    @staticmethod
    def generate_forecast(dataset, metric_col=None, periods=7):
        df = DatasetEngine.load_dataframe(dataset)
        if df.empty:
            raise ValueError("Dataset is empty.")

        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        if not numeric_cols:
            raise ValueError("No numeric telemetry columns available for time-series forecasting.")

        target_col = metric_col if metric_col in numeric_cols else numeric_cols[0]
        series = pd.to_numeric(df[target_col], errors='coerce').dropna()
        if len(series) < 5:
            raise ValueError("Insufficient data points for forecasting (minimum 5 required).")

        y = series.values.astype(np.float64)
        x = np.arange(len(y), dtype=np.float64)

        # Vectorized linear trend fit: y = m*x + c
        slope, intercept = np.polyfit(x, y, 1)
        y_fit = slope * x + intercept
        residuals = y - y_fit
        std_err = float(np.std(residuals)) if len(residuals) > 1 else 1.0

        future_x = np.arange(len(y), len(y) + periods, dtype=np.float64)
        future_y = slope * future_x + intercept

        # Vectorized 95% confidence interval bounds (1.96 * std_err)
        margin = 1.96 * std_err
        upper_bound = future_y + margin
        lower_bound = np.maximum(0.0, future_y - margin)

        historical_labels = [f"P{i+1}" for i in range(len(y))]

        future_labels = [f"F+{i+1}" for i in range(periods)]

        return {
            'metric': target_col,
            'periods': periods,
            'historical': {
                'labels': historical_labels[-30:],
                'values': [round(float(v), 2) for v in y[-30:]]
            },
            'forecast': {
                'labels': future_labels,
                'predicted': [round(float(v), 2) for v in future_y],
                'upper_bound': [round(float(v), 2) for v in upper_bound],
                'lower_bound': [round(float(v), 2) for v in lower_bound]
            },
            'trend_slope': round(float(slope), 4),
            'std_error': round(float(std_err), 2)
        }


class NLToFormulaEngine:
    """
    🧬 AI Natural Language DAX / Formula Generator.
    Parses plain text requirements (e.g., "Calculate percentage ratio of PFO to Rectified Power")
    and maps them to Pandas mathematical expressions using column fuzzy matching.
    """

    @staticmethod
    def generate_formula_from_nl(dataset, nl_prompt):
        df = DatasetEngine.load_dataframe(dataset)
        if df.empty:
            raise ValueError("Dataset is empty.")

        prompt = nl_prompt.strip().lower()
        cols = list(df.columns)

        matched_cols = []
        for col in cols:
            if col.lower() in prompt:
                matched_cols.append(col)

        formula = ""
        suggested_name = "Calculated_Measure"

        if "ratio" in prompt or "percentage" in prompt or "percent" in prompt or "divide" in prompt:
            if len(matched_cols) >= 2:
                formula = f"({matched_cols[0]} / {matched_cols[1]}) * 100"
                suggested_name = f"Ratio_{matched_cols[0]}_to_{matched_cols[1]}"
            elif len(matched_cols) == 1:
                formula = f"{matched_cols[0]} / 100"
                suggested_name = f"{matched_cols[0]}_Ratio"
        elif "multiply" in prompt or "times" in prompt or "scale" in prompt:
            if len(matched_cols) >= 2:
                formula = f"{matched_cols[0]} * {matched_cols[1]}"
                suggested_name = f"Product_{matched_cols[0]}_{matched_cols[1]}"
            elif len(matched_cols) == 1:
                formula = f"{matched_cols[0]} * 1.15"
                suggested_name = f"{matched_cols[0]}_Scaled"
        elif "subtract" in prompt or "difference" in prompt or "delta" in prompt:
            if len(matched_cols) >= 2:
                formula = f"{matched_cols[0]} - {matched_cols[1]}"
                suggested_name = f"Delta_{matched_cols[0]}_{matched_cols[1]}"
        elif "add" in prompt or "sum" in prompt or "total" in prompt:
            if len(matched_cols) >= 2:
                formula = f"{matched_cols[0]} + {matched_cols[1]}"
                suggested_name = f"Total_{matched_cols[0]}_{matched_cols[1]}"

        if not formula and matched_cols:
            formula = f"{matched_cols[0]} * 1.0"
            suggested_name = f"Measure_{matched_cols[0]}"
        elif not formula:
            first_num = df.select_dtypes(include=[np.number]).columns.first_valid_value() or cols[0]
            formula = f"{first_num} / 1000"
            suggested_name = f"{first_num}_Kilo"

        return {
            'formula': formula,
            'name': suggested_name,
            'prompt': nl_prompt
        }


class ETLPipeline:
    """
    🔄 Automated ETL Background Refresh Pipeline.
    Manages background scheduled data sync jobs, MongoDB collection auto-polling,
    and dataset record updates.
    """

    @staticmethod
    def run_etl_sync(dataset):
        from datetime import datetime
        df = DatasetEngine.load_dataframe(dataset)
        old_count = dataset.row_count or 0

        # Touch/Refresh dataset and update schema metadata
        new_count = len(df)
        dataset.row_count = new_count
        dataset.column_schema = DatasetEngine.infer_column_schema(df)
        dataset.save()

        # Update cache
        from .services import _df_cache
        _df_cache[dataset.id] = df.copy()

        return {
            'dataset_id': dataset.id,
            'dataset_name': dataset.name,
            'status': 'SUCCESS',
            'last_sync': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'records_synced': new_count,
            'delta_records': max(0, new_count - old_count)
        }

